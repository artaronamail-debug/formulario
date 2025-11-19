# pip install -r requirements.txt

import sys
import os
import json
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": ["null", "http://dantepropiedades.com.ar", "http://www.dantepropiedades.com.ar", "https://dantepropiedades.com.ar", "https://www.dantepropiedades.com.ar", "http://dantepropiedades.com", "https://danterealestate-github-io.onrender.com"]}})

EXCEL_FILE = 'contactos_dante_propiedades.xlsx'

def safe_print(message):
    safe_message = message.encode('ascii', 'ignore').decode('ascii')
    print(safe_message)

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"
        ws['A1'] = 'Fecha/Hora'
        ws['B1'] = 'Nombre'
        ws['C1'] = 'Firma'
        ws['D1'] = 'Teléfono'
        ws['E1'] = 'Propiedad'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        wb.save(EXCEL_FILE)
        safe_print(f"SUCCESS: Archivo {EXCEL_FILE} creado exitosamente")
    else:
        safe_print(f"INFO: Archivo {EXCEL_FILE} encontrado")

def serve_static_file(filename):
    try:
        if filename.startswith('api/') or filename in ['contactos_dante_propiedades.xlsx', 'propiedades.json']:
            safe_print(f"Archivo {filename} no se sirve como estático - será manejado por endpoint específico")
            return jsonify({"error": f"Archivo {filename} no encontrado"}), 404
        
        file_path = os.path.join(os.getcwd(), filename)
        safe_print(f"Buscando archivo: {file_path}")
        
        if os.path.exists(file_path):
            try:
                with open(file_path, 'rb') as f:
                    content = f.read()
                
                # Determinar el tipo de contenido
                if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.ico', '.svg')):
                    from flask import Response
                    return Response(content, mimetype=f'image/{filename.split(".")[-1].lower()}')
                elif filename.lower().endswith('.css'):
                    from flask import Response
                    return Response(content, mimetype='text/css')
                elif filename.lower().endswith('.js'):
                    from flask import Response
                    return Response(content, mimetype='application/javascript')
                elif filename.lower().endswith('.html'):
                    from flask import Response
                    return Response(content, mimetype='text/html')
                elif filename.lower().endswith('.json'):
                    from flask import Response
                    return Response(content, mimetype='application/json')
                else:
                    from flask import Response
                    return Response(content, mimetype='application/octet-stream')
                    
            except Exception as e:
                safe_print(f"Error leyendo {filename}: {str(e)}")
                return jsonify({"error": f"Error al leer archivo {filename}"}), 500
        else:
            safe_print(f"Archivo NO encontrado: {file_path}")
            return jsonify({"error": f"Archivo {filename} no encontrado"}), 404
            
    except Exception as e:
        safe_print(f"Error sirviendo {filename}: {str(e)}")
        return jsonify({"error": f"Error al servir {filename}"}), 500

def load_properties():
    try:
        with open('propiedades.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data
    except Exception as e:
        safe_print(f"Error cargando propiedades: {str(e)}")
        return []

@app.route('/')
def home():
    return serve_static_file('index.html')

@app.route('/api/guardar_contacto', methods=['POST'])
def guardar_contacto():
    try:
        data = request.get_json()
        
        # Validar que se recibieron los datos necesarios
        if not data or 'nombre' not in data:
            return jsonify({"error": "Datos incompletos"}), 400
        
        # Inicializar archivo Excel si no existe
        init_excel()
        
        # Cargar workbook existente
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Obtener siguiente fila
        next_row = ws.max_row + 1
        
        # Formatear fecha y hora
        now = datetime.now()
        fecha_hora = now.strftime("%Y-%m-%d %H:%M:%S")
        
        # Escribir datos
        ws[f'A{next_row}'] = fecha_hora
        ws[f'B{next_row}'] = data.get('nombre', '')
        ws[f'C{next_row}'] = data.get('email', '')
        ws[f'D{next_row}'] = data.get('telefono', '')
        ws[f'E{next_row}'] = data.get('propiedad_interes', '')
        
        # Guardar archivo
        wb.save(EXCEL_FILE)
        
        safe_print(f"Contacto guardado: {data.get('nombre')} - {data.get('email')}")
        
        return jsonify({
            "message": "Contacto guardado exitosamente",
            "fecha_hora": fecha_hora
        }), 200
        
    except Exception as e:
        safe_print(f"Error guardando contacto: {str(e)}")
        return jsonify({"error": f"Error al guardar contacto: {str(e)}"}), 500

@app.route('/api/properties', methods=['GET'])
def get_all_properties():
    try:
        properties = load_properties()
        return jsonify({
            "total": len(properties),
            "properties": properties
        }), 200
    except Exception as e:
        safe_print(f"Error obteniendo propiedades: {str(e)}")
        return jsonify({"error": f"Error al obtener propiedades"}), 500

def get_filter_options():
    try:
        properties = load_properties()
        barrios = []
        tipos = []
        
        for prop in properties:
            # **CORRECCIÓN CRÍTICA**: Validar tipo antes de usar .get()
            if not isinstance(prop, dict):
                safe_print(f"ADVERTENCIA: Elemento no es diccionario: {prop}")
                continue
            
            # Procesar barrio
            if prop.get('barrio') and prop['barrio'] not in barrios:
                barrios.append(prop['barrio'])
            
            # Procesar tipo
            if prop.get('tipo') and prop['tipo'] not in tipos:
                tipos.append(prop['tipo'])
        
        return {
            "barrios": sorted(barrios),
            "tipos": sorted(tipos),
            "total": len(properties)
        }
        
    except Exception as e:
        safe_print(f"Error en get_filter_options: {str(e)}")
        return {"error": f"Error obteniendo opciones de filtros: {str(e)}"}

@app.route('/api/properties/filter-options', methods=['GET'])
def get_filter_options_endpoint():
    try:
        options = get_filter_options()
        if "error" in options:
            return jsonify(options), 500
        
        return jsonify(options), 200
        
    except Exception as e:
        safe_print(f"Error en endpoint filter-options: {str(e)}")
        return jsonify({"error": f"Error en servidor: {str(e)}"}), 500

def search_properties(ope=None, tipo=None, loc=None, precio_max=None, ambientes=None):
    try:
        properties = load_properties()
        results = []
        
        for prop in properties:
            # **CORRECCIÓN CRÍTICA**: Validar tipo antes de usar .get()
            if not isinstance(prop, dict):
                safe_print(f"ADVERTENCIA: Elemento no es diccionario en search: {prop}")
                continue
            
            # Aplicar filtros
            if ope and prop.get('operacion') != ope:
                continue
                
            if tipo and prop.get('tipo') != tipo:
                continue
                
            if loc and prop.get('barrio', '').lower() != loc.lower():
                continue
                
            if precio_max:
                try:
                    prop_precio = float(prop.get('precio', 0))
                    if prop_precio > precio_max:
                        continue
                except (ValueError, TypeError):
                    continue
                    
            if ambientes:
                try:
                    prop_amb = int(prop.get('ambientes', 0))
                    if prop_amb < ambientes:
                        continue
                except (ValueError, TypeError):
                    continue
            
            results.append(prop)
        
        return results
        
    except Exception as e:
        safe_print(f"Error en search_properties: {str(e)}")
        return []

@app.route('/api/properties/search', methods=['GET'])
def search_properties_endpoint():
    try:
        # Obtener parámetros de la URL
        ope = request.args.get('ope')
        tipo = request.args.get('tipo')
        loc = request.args.get('loc')
        precio_max = request.args.get('precio_max')
        ambientes = request.args.get('ambientes')
        
        # Convertir tipos de datos
        if precio_max:
            try:
                precio_max = float(precio_max)
            except ValueError:
                precio_max = None
                
        if ambientes:
            try:
                ambientes = int(ambientes)
            except ValueError:
                ambientes = None
        
        safe_print(f"--- Nueva Búsqueda ---")
        safe_print(f"Parámetros recibidos: ope={ope}, tipo={tipo}, loc={loc}, cod={request.args.get('cod')}, precio_min={request.args.get('precio_min')}, precio_max={precio_max}, ambientes={ambientes}")
        
        results = search_properties(ope, tipo, loc, precio_max, ambientes)
        
        safe_print(f"Propiedades encontradas: {len(results)}")
        
        return jsonify({
            "total": len(results),
            "properties": results,
            "filters": {
                "operacion": ope,
                "tipo": tipo,
                "localidad": loc,
                "precio_max": precio_max,
                "ambientes_min": ambientes
            }
        }), 200
        
    except Exception as e:
        safe_print(f"Error en endpoint search: {str(e)}")
        return jsonify({"error": f"Error en servidor: {str(e)}"}), 500

@app.route('/api/properties/stats', methods=['GET'])
def get_stats():
    try:
        properties = load_properties()
        
        total_propiedades = len(properties)
        tipos_count = {}
        barrios_count = {}
        operaciones_count = {}
        
        for prop in properties:
            # **CORRECCIÓN CRÍTICA**: Validar tipo antes de usar .get()
            if not isinstance(prop, dict):
                safe_print(f"ADVERTENCIA: Elemento no es diccionario en stats: {prop}")
                continue
            
            # Contar tipos
            tipo = prop.get('tipo', 'Sin especificar')
            tipos_count[tipo] = tipos_count.get(tipo, 0) + 1
            
            # Contar barrios
            barrio = prop.get('barrio', 'Sin especificar')
            barrios_count[barrio] = barrios_count.get(barrio, 0) + 1
            
            # Contar operaciones
            operacion = prop.get('operacion', 'Sin especificar')
            operaciones_count[operacion] = operaciones_count.get(operacion, 0) + 1
        
        return jsonify({
            "total_propiedades": total_propiedades,
            "tipos_mas_comunes": dict(sorted(tipos_count.items(), key=lambda x: x[1], reverse=True)[:10]),
            "barrios_mas_comunes": dict(sorted(barrios_count.items(), key=lambda x: x[1], reverse=True)[:10]),
            "operaciones": operaciones_count
        }), 200
        
    except Exception as e:
        safe_print(f"Error obteniendo stats: {str(e)}")
        return jsonify({"error": f"Error al obtener estadísticas: {str(e)}"}), 500

# **RUTA GENÉRICA AL FINAL** - Debe ser la última
@app.route('/<path:filename>')
def serve_any_file(filename):
    return serve_static_file(filename)

if __name__ == '__main__':
    safe_print("Iniciando servidor Flask...")
    init_excel()
    safe_print("Servidor listo para recibir solicitudes")
    
    # Usar puerto 10000 para Render, 5000 para desarrollo local
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)